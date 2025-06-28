import openpyxl
import csv
import os
import logging
from datetime import datetime, date
from project_logger import setup_project_logger
from config import RAW_DATA_DIR, PROCESSED_DATA_DIR

logger = setup_project_logger("excel_extraction") # Initialize the logger with a module-specific name

def _clean_text(text):
    """
    Cleans common typographical characters like smart quotes and dashes.
    Replaces them with their ASCII equivalents.
    """
    if text is None:
        return None
    text = str(text)
    text = text.replace('’', "'") # Right single quotation mark
    text = text.replace('‘', "'") # Left single quotation mark
    text = text.replace('“', '"') # Left double quotation mark
    text = text.replace('”', '"') # Right double quotation mark
    text = text.replace('–', '-') # En dash
    text = text.replace('—', '--') # Em dash (longer dash)
    text = text.replace(' ', ' ') # Non-breaking space to regular space
    text = text.replace('\n', ' ') # Remove newlines
    text = text.replace('\r', '') # Remove carriage returns
    text = text.replace('\t', ' ') # Replace tabs with spaces
    text = text.replace('â€', "'") # Common mojibake for apostrophe/right single quote
    text = text.replace('â€œ', '"') # Common mojibake for left double quote
    text = text.replace('â€�', '"') # Common mojibake for right double quote
    text = text.replace('â€“', '-') # Common mojibake for en dash
    text = text.replace('â€”', '--') # Common mojibake for em dash
    text = text.replace('…', '...') # Ellipsis to three dots
    # Add more replacements here if other specific characters are found
    return text.strip()

def extract_data_from_single_sheet(workbook, sheet_name):
    """
    Extracts article titles, URLs, and source websites from a single sheet
    in an Excel workbook, also identifying if a title cell is highlighted.
    Applies text cleaning to titles and URLs.

    Args:
        workbook (openpyxl.workbook.workbook.Workbook): The loaded openpyxl workbook object.
        sheet_name (str): The name of the sheet to process.

    Returns:
        list: A list of dictionaries, where each dictionary represents an article
              and contains 'source_website', 'title', 'url', 'is_selected',
              and 'origin_sheet'.
    """
    sheet = workbook[sheet_name]
    articles = []
    header_row_values = [cell.value for cell in sheet[1]]

    # Iterate through columns in steps of 2 (A-B, C-D, etc.)
    for col_idx in range(0, len(header_row_values), 2):
        source_website = header_row_values[col_idx]
        if source_website is None:
            # If the source website cell is empty, this column pair is likely unused
            logger.debug(f"Skipping empty source website column at index {col_idx}.")
            continue

        # Iterate through rows starting from the second row
        for row_idx in range(2, sheet.max_row + 1):
            title_cell = sheet.cell(row=row_idx, column=col_idx + 1)
            url_cell = sheet.cell(row=row_idx, column=col_idx + 2)

            title = _clean_text(title_cell.value) # Apply cleaning here
            url = _clean_text(url_cell.value)     # Apply cleaning here

            if title is None and url is None:
                # If both title and URL are empty, assume no more articles in this column
                logger.debug(f"No more data in column for source '{source_website}' from row {row_idx}. Breaking.")
                break

            is_selected = "Not Selected"
            if title_cell.fill and title_cell.fill.patternType and title_cell.fill.fgColor.rgb:
                # openpyxl returns '00000000' for no fill or solid white by default
                # We are looking for any non-transparent fill that's not pure black (often default for empty)
                if title_cell.fill.fgColor.rgb != '00000000' and title_cell.fill.fgColor.rgb != 'FFFFFFFF':
                    is_selected = "Selected"
                    logger.debug(f"Title '{title}' for '{source_website}' at row {row_idx} is selected.")

            if title or url: # Only add if there's at least a title or URL
                articles.append({
                    'source_website': source_website,
                    'title': title,
                    'url': url,
                    'selected': is_selected,
                    'date': sheet_name  # Add the origin sheet name as the 'date' for the record
                })
                logger.debug(f"Extracted: '{title}' from '{source_website}' on sheet '{sheet_name}'.")
    return articles

def process_all_sheets_and_write_to_csv(excel_file_path, output_csv_file_path):
    """
    Processes all sheets in an Excel file, extracts data, and writes
    the consolidated data to a CSV file.

    Args:
        excel_file_path (str): The path to the input Excel (.xlsx) file.
        output_csv_file_path (str): The path to the output CSV file.
    """
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        logger.error(f"Error: The Excel file '{excel_file_path}' was not found.")
        return
    except Exception as e:
        logger.error(f"Error loading workbook '{excel_file_path}': {e}", exc_info=True)
        return

    all_extracted_data = []
    logger.info(f"Processing Excel file: {excel_file_path}")
    logger.info(f"Found {len(workbook.sheetnames)} sheets: {', '.join(workbook.sheetnames)}")

    for sheet_name in workbook.sheetnames:
        logger.info(f"  Extracting data from sheet: '{sheet_name}'...")
        sheet_data = extract_data_from_single_sheet(workbook, sheet_name)
        all_extracted_data.extend(sheet_data)
        logger.info(f"    Extracted {len(sheet_data)} entries from '{sheet_name}'.")

    if all_extracted_data:
        # Define the headers for your CSV file, including the new 'date'
        csv_fieldnames = ['source_website', 'title', 'url', 'selected', 'date']
        try:
            # Ensure the output directory exists
            output_dir = os.path.dirname(output_csv_file_path)
            os.makedirs(output_dir, exist_ok=True)
            
            file_exists = os.path.isfile(output_csv_file_path)
            
            with open(output_csv_file_path, 'a', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=csv_fieldnames)
                if not file_exists:
                    writer.writeheader() # Write the header row
                writer.writerows(all_extracted_data) # Write all the data rows
            logger.info(f"\nSuccessfully consolidated data from all sheets and wrote to '{output_csv_file_path}'")
            logger.info(f"Total entries written: {len(all_extracted_data)}")
            logger.info(f"Total sheets read: {len(workbook.sheetnames)}")
        except Exception as e:
            logger.error(f"Error writing consolidated data to CSV file '{output_csv_file_path}': {e}", exc_info=True)
    else:
        logger.warning("\nNo data extracted from any sheet to write to CSV.")

def process_files_in_directory(daily_trackers_folder=RAW_DATA_DIR, processed_data_folder=PROCESSED_DATA_DIR):
    """
    Processes all Excel files in the specified directory and writes
    the extracted data to corresponding CSV files.

    Args:
        daily_trackers_folder (str): The path to the folder containing daily tracker Excel files.
        processed_data_folder (str): The path to the folder where output CSV files will be saved.
    """
    # Ensure the output directory exists
    os.makedirs(processed_data_folder, exist_ok=True)
    logger.info(f"Ensured output directory '{processed_data_folder}' exists.")

    # List all files in the daily_trackers folder
    excel_files = [f for f in os.listdir(daily_trackers_folder) if f.endswith('.xlsx')]

    if not excel_files:
        logger.warning(f"No Excel files found in '{daily_trackers_folder}'. Please check the path and file extensions.")
        return

    today_date = date.today()
    for excel_file_name in excel_files:
        excel_file_path = os.path.join(daily_trackers_folder, excel_file_name)

        output_csv_file_name = f"processed_{today_date.strftime('%Y%m%d')}.csv"
        output_csv_file_path = os.path.join(processed_data_folder, output_csv_file_name)

        logger.info(f"\n--- Processing '{excel_file_name}' ---")
        process_all_sheets_and_write_to_csv(excel_file_path, output_csv_file_path)
        logger.info(f"--- Finished processing '{excel_file_name}' ---\n")
        
    logger.info("All Excel files processed successfully.")
    return True
        
        
# --- How to use the code ---
if __name__ == "__main__":
    process_files_in_directory()